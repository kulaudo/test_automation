from openpyxl import load_workbook

class ParseExcel:
    def __init__(self, excel_path, sheet_name):
        self.wb = load_workbook(excel_path)
        self.sheet = self.wb[sheet_name]


    def get_data_from_sheet(self):
        data_list = []
        for cols in self.sheet.iter_rows(min_row=2):
            temp_list = []
            for cell in cols:            
                temp_list.append(cell.value)
            data_list.append(temp_list)
        return data_list
    
if __name__ == "__main__":
    excel_path = 'E:\\test_automation\\test\\test_data\\test_sheet.xlsx'
    sheet_name = 'search_data'
    pe = ParseExcel(excel_path=excel_path, sheet_name=sheet_name)
    for cols in pe.get_data_from_sheet():
        for item in cols:
            print(item, end=' ')
        print()
        



        